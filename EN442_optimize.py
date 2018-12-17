# -*- coding: utf-8 -*-
"""
Created on Mon Aug 13 14:42:36 2018

@author: ondraZ

MIT License

Copyright (c) [2018] Ondrej Zlevor, ondrej.zlevor@cvut.cz

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

Any modifications of this source code shall be documented and reported to 
the author, preferably by a pull request to the repository, in order to contribute 
to the public open source. 

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

"""

import numpy as np
import math
from scipy.optimize import minimize
from scipy.optimize import show_options
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import openpyxl
import os
from os import system, name 
import sys

import time

TEMPERATURE_DIFF_COEFF = 2
C_P = 4185
MAX_Q_N_RATIO = 2
INIT_ROWS = 2

def objective(x, args):
    """
    objective(x, args) - cost function for optimization
    INPUTS:
        x:      I)   [q_n, m_dot]
                II)  [m_dot, m_dot_p]
                III) [m_dot]
    
        args:   q_d         - desired heat flows [W]
                Twi         - inlet water temperature [degC]
                Two_d       - desired return water temperature [degC]
                Tai         - surrounding air temperature [degC]
                LMTD_n      - LMTD value computed from catalogue nominal values [degC]
                scaleFactor - scale factor of variables [-]
                tag         - tag describing type of simulation run [str]
                (bnds)      - optimization variables boundaries (scaled) [-]
                (q_n)       - HX nominal heat flows [W]
                (m_dot_p)   - primary loop flow (scaled) [kg/s]
            
    """

    W_FOUT = 1000   # weight of penalization of non-correct actual heat flow
    W_Q_N = 0.0001  # weight of penalization of HX size
    W_M_DOT = 0.2
#    W_M_DOT = 0
#    W_Q_N = 0
    

    q_d, Twi, Two_d, Tai, LMTD_n,scFactor,n_coeff = args['q_d'],args['Twi'],\
        args['Two'],args['Tai'],args['LMTD_n'], args['scaleFactor'], args['n_coeff']

    x = np.multiply(x,scFactor[0:len(x)])
    n = len(q_d);    
    m_dot = x[0:n]
    
    # === Choose type of simulation run ===
    if len(x) > n+1:
        # --- x = [q_n, m_dot] ---
        q_n = x[n:len(x)]
        m_dot_p = np.sum(q_d)/4185/(Twi-Two_d)
        bnds = args['bnds']
    elif len(x) > n:
        # --- x = [m_dot, m_dot_p] ---
        q_n = args['q_n']
        m_dot_p = x[n]
    else:
        # --- x = [m_dot] ---
        q_n = args['q_n']
        m_dot_p = args['m_dot_p']
        
        
    
    # === Compute actual heat flows from the HX ===
    
    fout = np.zeros(n)
    Twi_vec = np.zeros(n)
    Two_vec = np.zeros(n)
    LMTD_vec = np.zeros(n)
    obj_func_penalty = 0
    
    for iHX in range(0,n):
        if (Twi-Tai-q_d[iHX]/4185/m_dot[iHX]) < 0 or any(m_dot > np.ones(n)*m_dot_p):
            if (Twi-Tai-q_d[iHX]/4185/m_dot[iHX]) < 0:
                m_dot_old = m_dot[iHX].copy()
                m_dot[iHX] = q_d[iHX]/(Twi-Tai)/4185*1.001
                obj_func_penalty = ((m_dot_old - m_dot[iHX])*2e5)**2
            else:
                 return 1e5
        LMTD = (q_d[iHX]/4185/m_dot[iHX])/math.log((Twi-Tai)/(Twi-Tai-q_d[iHX]/4185/m_dot[iHX]))
        fout[iHX] = math.pow((q_d[iHX]/q_n[iHX] - (LMTD/LMTD_n[iHX])**n_coeff),2)
        Two = Twi - q_d[iHX]/m_dot[iHX]/4185
        Twi = Twi - m_dot[iHX]/m_dot_p*(Twi-Two)
        Twi_vec[iHX] = Twi
        Two_vec[iHX] = Two        
        LMTD_vec[iHX] = LMTD
        
    # === Return results ===    
        
    if args['tag'] is 'opt':
#        return np.sum(fout)*W_FOUT + np.sum(q_n)*W_Q_N + np.sum(m_dot)
        bnds_dist = -0.95*bnds[0][1] + max(m_dot)
        return ((np.sum(fout))*W_FOUT) + np.sum(q_n)*W_Q_N + ((1+np.sign(bnds_dist))*bnds_dist*5)**2 + np.sum(m_dot)*W_M_DOT
#        return np.sum(fout)
    elif args['tag'] is 'analyze':
        bnds_dist = -0.95*bnds[0][1] + max(m_dot)
        return {'w_fout': ((np.sum(fout))*W_FOUT), 'w_q_n': np.sum(q_n)*W_Q_N, 'bnds': ((1+np.sign(bnds_dist))*bnds_dist*5)**2 }
    elif args['tag'] is 'real':
#        return np.sum(fout)*W_FOUT + (Twi-Two_d)**2
        m_dot_dist = -0.95*m_dot_p + max(m_dot)
#        return np.sum(fout)  + ((1+np.sign(m_dot_dist))*m_dot_dist*100)**2 + obj_func_penalty
        return np.sum(fout)
    elif args['tag'] is 'analyze_real':
        m_dot_dist = -0.95*m_dot_p + max(m_dot)
        return {'w_fout': ((np.sum(fout))*W_FOUT), 'w_m_dot': ((1+np.sign(m_dot_dist))*m_dot_dist*100)**2 }
    elif args['tag'] is 'sim':
        return {'Twi':Twi_vec, 'Two': Two_vec, 'LMTD': LMTD_vec, 'fout': fout}

    
def find_minimal_solution(x0,args):
    """
    find_minimal_solution(x0,args) - finds minimal m_dot_p for that given HX set has a solution
    INPUTS:
        x0      - initial m_dot
        
        args:   q_d         - desired heat flows [W]
                Twi         - inlet water temperature [degC]
                Two_d       - desired return water temperature [degC]
                Tai         - surrounding air temperature [degC]
                LMTD_n      - LMTD value computed from catalogue nominal values [degC]
                scaleFactor - scale factor of variables [-]
                bnds        - optimization variables boundaries (scaled) [-]
    """
    MIN_M_DOT_RATIO = 64
    FUN_TOLERANCE = 1e-6
    
    
    q_d, Twi, Two_d, bnds= args['q_d'],args['Twi'], args['Two'],args['bnds']
    m_dot_p_base = np.sum(q_d)/4185/(Twi-Two_d)
    solution_vec = ()
    m_dot_vec = ()
    
    m_dot_p = m_dot_p_base
    m_dot_step = m_dot_p_base
    
    args['m_dot_p'] = m_dot_p
    solution = minimize(fun = objective, args = args, x0 = x0,\
                method='SLSQP', options={'ftol': 1e-17, 'disp': True,'maxiter': 60, 'iprint': 0},\
                bounds = bnds[0:n])  
    if solution.fun < 2e-6:
        stopTag = True
    else:
        stopTag = False
        m_dot_p = m_dot_p + m_dot_step
    
    min_sol = solution
    min_sol_m_dot_p = m_dot_p
    
    # === Search minimal m_dot_p, for that system has solution ===
    while not stopTag:
        
        args['m_dot_p'] = m_dot_p
        solution = minimize(fun = objective, args = args, x0 = x0,\
                    method='SLSQP', options={'ftol': 1e-9, 'disp': True,'maxiter': 60,  'iprint': 0},\
                    bounds = bnds[0:n]) 
        solution_vec = solution_vec + (solution,)
        m_dot_vec = m_dot_vec + (m_dot_p,)
        
        # --- save solution with minimal m_dot_p ---
        if solution.fun < min_sol.fun:
            min_sol = solution
            min_sol_m_dot_p = m_dot_p
        
        if m_dot_step < m_dot_p_base/MIN_M_DOT_RATIO:
            return {'solution': min_sol, 'm_dot_p': min_sol_m_dot_p,'solution_vec': solution_vec, 'm_dot_vec': m_dot_vec}
        
        if solution.fun < FUN_TOLERANCE:
            m_dot_step = m_dot_step/2
            m_dot_p = m_dot_p - m_dot_step
        else:
            if m_dot_step < m_dot_p_base:
                m_dot_step = m_dot_step/2
            m_dot_p = m_dot_p + m_dot_step
            if m_dot_p > 5*m_dot_p_base:
                return {'solution_vec': solution_vec, 'm_dot_vec': m_dot_vec}
    return {'solution': solution, 'm_dot_p': m_dot_p,}


def constraint2(x):
    n = len(x) - 1
    if any(x[0:n] > np.ones(n)*x[n]):
        return -1
    else:
        return 1

def readRangeAsVector(sheet,cellStart, cellEnd):
    """
    readRangeAsVector(sheet,cellStart, cellEnd) - read values range in a *.xls 
        file and save them into a list (vector)
    INPUTS:
        sheet       - xls sheet object reference
        cellStart   - start cell of range [str]
        cellEnd     - end cell of range [str]
    """
    if cellStart[0] == cellEnd[0]:
        if type(sheet[cellStart].value) == float or type(sheet[cellStart].value) == int:
            output_vec = np.zeros(int(cellEnd[1:len(cellEnd)]) - int(cellStart[1:len(cellStart)])+1)
            k = 0
            for cellObj in sheet[cellStart:cellEnd]:
                output_vec[k] = cellObj[0].value
                k = k+1
            return output_vec[~np.isnan(output_vec)]
        else:
            output_vec = ()
            k = 0
            for cellObj in sheet[cellStart:cellEnd]:
                output_vec = output_vec + (cellObj[0].value,)
                k = k+1       
            return output_vec
    
def buildHeap(q_chain, n):
    """
    buildHeap(q_chain, n) - recursive method to generate all permutations of 
        0,1 of length n
    """
    if len(q_chain) < n:
        q_chain_part0 = buildHeap(q_chain + [0],n)
        q_chain_part1 = buildHeap(q_chain + [1],n)
        
        if isinstance(q_chain_part0,tuple):
            returnVal = q_chain_part0 + q_chain_part1
        else:
            returnVal = ((q_chain_part0) , (q_chain_part1))
    else:
        returnVal = (q_chain)
    return returnVal

def fillRange(ws, cellStart, cellEnd, colCode):
    """
    fillRange(ws, cellStart, cellEnd, colCode) - fill range in an *.xls file 
        by given color
    INPUTS:
        ws          - worksheet reference
        cellStart   - range starting cell (rectangle left top)
        cellEnd     - range ending cell (rectangle right bottom)
        colCode     - color code
    """
    cellObj = sheet[cellStart:cellEnd][0]
    for iCell in range(0,len(cellObj)):
        cellObj[iCell].fill =  PatternFill(fill_type="solid",
                   start_color=colCode,
                   end_color=colCode)

# define our clear function 
def clear(): 
    """
    clear() - clears terminal window
    """
  
    # for windows 
    if name == 'nt': 
        _ = system('cls') 
  
    # for mac and linux(here, os.name is 'posix') 
    else: 
        _ = system('clear') 



print('Number of arguments:' + str(len(sys.argv)) + 'arguments.')
print('Argument List:' + str(sys.argv))

start_time1 = time.time()

# Change directory 
dir_path = os.path.join(os.getcwd(),'../../')
print('Current directory :' + dir_path)
dir_path = os.path.realpath(dir_path)

if len(sys.argv) > 1:
    dir_path = sys.argv[1]
    xls_name = sys.argv[2]
else:
    dir_path = os.getcwd()
    xls_name = 'Onepipe_design_setup.xlsm'
    

xls_name_res = xls_name.replace('setup','result')
xls_name_res = xls_name_res.replace('_CZ','')
xls_name_res = xls_name_res.replace('xlsm','xlsx')


print('Current directory :' + dir_path)
print('xls_name :' + xls_name)

os.chdir(dir_path)

    

wb = openpyxl.load_workbook(dir_path + '/' + xls_name,data_only=True)

# ----- Get User-set values ------
sheet = wb.get_sheet_by_name('Project_setup')

problem_setup = {'Twi': sheet['F3'].value, 'Two': sheet['F4'].value,\
                 'Tai': sheet['F5'].value, 'm_dot_max_1': sheet['F6'].value,\
                 'm_dot_max_2': sheet['F7'].value,\
                 'catalogue': sheet['AE' + str(sheet['AF3'].value + 2)].value,\
                 'q_d': readRangeAsVector(sheet,'F10','F29'),\
                 'LMTD_n' : ([50, 50, 50]),'tag': 'opt', 'n_coeff':sheet['F7'].value}



# initial guesses
n = len(problem_setup['q_d'])
sheet_cat = wb.get_sheet_by_name(problem_setup['catalogue'])
problem_setup['LMTD_n'] = sheet_cat['I3'].value*np.ones(n)

LMTD_actual = (problem_setup['Twi'] - problem_setup['Two'])/\
np.log((problem_setup['Twi'] - problem_setup['Tai'])/(problem_setup['Two'] - problem_setup['Tai']))
problem_setup['LMTD_n'] = LMTD_actual*np.ones(n)

q_n_init_scale = (sheet_cat['I3'].value/LMTD_actual)**1.3

x0 = np.concatenate((problem_setup['q_d']/C_P/\
                     (problem_setup['Twi'] - problem_setup['Two'])*\
                     TEMPERATURE_DIFF_COEFF, problem_setup['q_d'])\
    ,axis=0)
    
# Scale Factors
scFactor = np.concatenate((np.ones(n)*max(x0[0:n]), np.ones(n)*max(problem_setup['q_d'])), axis = 0 )
x0 = np.divide(x0, scFactor)
problem_setup['scaleFactor'] = scFactor


# optimize
m_dot_min_vec = problem_setup['q_d']/C_P/(problem_setup['Two'] - problem_setup['Tai'])*1.0001
m_dot_max_vec = np.sum(problem_setup['q_d'])/C_P/(problem_setup['Twi'] - problem_setup['Two'])*np.ones(n)

b_m_dot = np.array([m_dot_min_vec,m_dot_max_vec])
b_m_dot = b_m_dot/scFactor[0]
b_m_dot = b_m_dot.transpose()

b_q = (0,max(problem_setup['q_d'])*MAX_Q_N_RATIO)
b_q = b_q/scFactor[n]

bnds = (tuple(map(tuple, b_m_dot))) + (b_q,)*n
problem_setup['bnds'] = bnds

cons = {'type':'eq', 'fun': constraint2}





# One step optimization

q_heap = buildHeap([], n)
all_sol = ()
all_fun = np.zeros(len(q_heap))
status_vec = np.zeros(len(q_heap))
for iOpt in range(0,len(q_heap)):
    w_vec = np.ones(n)*0.7 + np.array(q_heap[iOpt])*0.6
    x0set = x0
    x0set[n:2*n] = np.multiply(x0[n:2*n],w_vec)

    solution = minimize(fun = objective,\
                        args = problem_setup,\
                        x0 = x0set, method='SLSQP',\
                        options={'ftol': 1e-12, 'disp': True, 'iprint': 1},\
                        bounds = bnds)
    status_vec[iOpt] = solution.status
    all_sol = all_sol + (solution,)
    all_fun[iOpt] = solution.fun

success_sol = np.extract(status_vec == 0, all_sol)
if not any(success_sol):
    success_sol = np.extract(status_vec == 9, all_sol)
        
    
if any(success_sol):
    fun_vals = [s.fun for s in success_sol]

    solution = success_sol[np.argmin(fun_vals)] 
    solution.x

    
    x = np.multiply(solution.x,scFactor)
    q_n = x[n:len(x)]
    
    # Get real heat flow
    problem_setup['tag'] = 'sim'
    temps = objective(solution.x,problem_setup)
    Twi_vec_opt = temps['Twi'].copy()
    Twi_vec = temps['Twi']
    Two_vec = temps['Two']
    Twi_vec[1:n] = Twi_vec[0:n-1]
    Twi_vec[0] = problem_setup['Twi']
    LMTD_vec = temps['LMTD']
    Q_opt = np.multiply((LMTD_vec/problem_setup['LMTD_n'])**problem_setup['n_coeff'], q_n)
    m_dot_opt= x[0:n]
    q_n_opt = q_n
    m_dot_p_opt = np.sum(problem_setup['q_d'])/4185/(problem_setup['Twi']-problem_setup['Two'])
#    ---- Pokus ----
    q_n_init_scale = (sheet_cat['I3'].value/problem_setup['LMTD_n'])**problem_setup['n_coeff']
#    q_n_opt = q_n_init_scale*q_n
    q_n = q_n_init_scale*q_n
    q_n_sc = q_n
    problem_setup['LMTD_n'] = sheet_cat['I3'].value*np.ones(n)
    
    x = solution.x
    x[n:2*n] = x[n:2*n]*q_n_init_scale
    xr = np.multiply(x,scFactor)
    q_n = xr[n:len(xr)]
    temps = objective(x,problem_setup)
    LMTD_vec = temps['LMTD']
    Q_opt_sc = np.multiply((LMTD_vec/problem_setup['LMTD_n'])**problem_setup['n_coeff'], q_n)
#    --------------
    Two_vec_opt = Two_vec
    
    start_time2 = time.time()



#%%

# ------------------------------------------------------------------------
# =============================================================================
# # ------------------------- Find real units ------------------------------
# =============================================================================
# ------------------------------------------------------------------------


sheet = wb.get_sheet_by_name(problem_setup['catalogue'])

q_n_unsrt = readRangeAsVector(sheet,'F3','F'+str(2+sheet['N3'].value) )
q_n_asc = np.sort(q_n_unsrt)
q_n_names = readRangeAsVector(sheet,'K3','K'+str(2+sheet['N3'].value) )

problem_setup_m = problem_setup


q_heap = buildHeap([], n)

q_n_higher = np.zeros(n)
q_n_lower = np.zeros(n)
q_n_higher_lower = np.zeros((n,2))
solution_vec = ()
Q_res = ()
Q_max_vec = ()
Two_all = np.zeros((len(q_heap),n))
Twi_all = np.zeros((len(q_heap),n))
m_dot_all = np.zeros((len(q_heap),n+1))
q_n_all = np.zeros((len(q_heap),n))
fout_all = np.zeros((len(q_heap),n))
Q_res = np.zeros((len(q_heap),n))

exec_time_vec = np.zeros(len(q_heap))

final_obj = np.zeros(len(q_heap))

if any(success_sol):
# --- If there is an optimal solution, go through solutions with real HX ---

    for iHX in range(0,len(q_n)):
        q_n_higher_vec = q_n_asc[q_n_asc > q_n[iHX]] if any(q_n_asc > q_n[iHX]) else [q_n_asc[len(q_n_asc)-1]]
        q_n_lower_vec = q_n_asc[q_n_asc < q_n[iHX]] if any(q_n_asc < q_n[iHX]) else [q_n_asc[0]]
        q_n_higher_lower[iHX][0] = q_n_higher_vec[0]    
        q_n_higher_lower[iHX][1] = q_n_lower_vec[len(q_n_lower_vec)-1]
        
    for iSet in range(0,len(q_heap)):
        
        clear()
        print('Solution ' + str(iSet) + '/' + str(len(q_heap)))
        print('||' + '='*round(iSet/len(q_heap)*100) + '_' * (100 - round(iSet/len(q_heap)*100)) + '||' )
        last_time = time.time()
        q_n_set = np.zeros(n)
        for iHX in range(0,len(q_heap[iSet])):
            q_n_set[iHX] = q_n_higher_lower[iHX][q_heap[iSet][iHX]]
            
        problem_setup_m['q_n'] = q_n_set
        problem_setup_m['tag'] = 'real'
        
        if (np.multiply(q_heap[iSet],q_n_higher_lower.transpose()[1]) < q_n_higher_lower.transpose()[0]).all():
            x0 = np.append(m_dot_opt,np.sum(problem_setup_m['q_d'])/C_P/(problem_setup_m['Twi'] - problem_setup_m['Two']))
            scFactor = np.ones(n+1)*max(x0[0:n])
            x0 = np.divide(x0, scFactor)
            problem_setup_m['scaleFactor'] = scFactor
            
            
            m_dot_min_vec = problem_setup_m['q_d']/C_P/(problem_setup_m['Two'] - problem_setup_m['Tai'])*1.1
            m_dot_max_vec = np.sum(problem_setup_m['q_d'])/C_P/((problem_setup_m['Twi'] - problem_setup_m['Two'])/n*2)*np.ones(n)
            b_m_dot = np.array([m_dot_min_vec*0,m_dot_max_vec])
            b_m_dot = b_m_dot/scFactor[0]
            b_m_dot = b_m_dot.transpose()
     
            bnds =  (tuple(map(tuple, b_m_dot)))
            
            
            
            bnds = bnds + ((max(m_dot_min_vec/scFactor[0]),1/scFactor[0]),)
            cons = ({'type': 'ineq', 'fun':constraint2})
    
            problem_setup_m['x0'] = x0[0:n]
            problem_setup_m['bnds'] = bnds
            
    
            sol_vec = find_minimal_solution(x0[0:n],problem_setup_m)
            if 'solution' in sol_vec:
                solution = sol_vec['solution']
                solution.x = np.append(solution.x,sol_vec['m_dot_p']/scFactor[n])
                problem_setup_m['scaleFactor'] = scFactor
        
            
            
                solution_vec = solution_vec + (solution,)
                
                
                problem_setup_m['tag'] = 'sim'
                temps = objective(solution.x,problem_setup_m)
                Twi_vec = temps['Twi']
                Two_vec = temps['Two']
                Twi_all[iSet][:] = Twi_vec
                Twi_vec[1:n] = Twi_vec[0:n-1]
                Twi_vec[0] = problem_setup['Twi']
                Two_all[iSet][:] = Two_vec
                m_dot_all[iSet][:] = np.multiply(scFactor,solution.x)
                fout_all[iSet]= temps['fout']
                LMTD_vec = temps['LMTD']
                Q_max = np.multiply(np.power(([xHX - problem_setup_m['Tai'] for xHX in Twi_vec])\
                    /problem_setup_m['LMTD_n'],1.3),q_n_set)
                Q = np.multiply((LMTD_vec/problem_setup_m['LMTD_n'])**problem_setup_m['n_coeff'], q_n_set)
      
                Q_res[iSet][:] = Q
                Q_max_vec = Q_max_vec + (Q_max,)
                final_obj[iSet] = np.sum(np.power((problem_setup['q_d'] - Q),2))
        q_n_all[iSet][:] = q_n_set
        exec_time_vec[iSet] = time.time() - last_time
    
    
    indObj = [i for i in range(0,len(final_obj)) if final_obj[i]<1]
    
    
    stop_time = time.time()
    
    print("--- Execution time ---")
    print("Loop 1: %s seconds" % (start_time2 - start_time1))
    print("Loop 2: %s seconds" % (stop_time - start_time2))
    print("Whole program: %s seconds" % (stop_time - start_time1))

#%%

# ------------ Open result sheet -------------------

wb.close()
varSpace = 4

try:
    wb = openpyxl.load_workbook('./' + xls_name_res,data_only=True)
    tag = 'exists'
except:
    wb = Workbook(xls_name_res.replace('.xlsx',''))
    wb.save(xls_name_res)
    wb.close
    wb = openpyxl.load_workbook('./' + xls_name_res, data_only=True)
    tag = 'non-exists'
    


# ------------------------------------------------------------------------
# ----------------------- Write optimal results to sheet -------------------------
# ------------------------------------------------------------------------
    
sh_names = wb.get_sheet_names()
if any("Optimal_results" == s for s in sh_names):  
    sheet = wb.get_sheet_by_name('Optimal_results')
    wb.remove_sheet(sheet)
sheet = wb.create_sheet('Optimal_results')
    
if any(success_sol):    
    sheet.cell(row=1,column=1,value='Fobj')
    sheet.cell(row=1,column=2,value=final_obj[iSet])
    sheet.cell(row=1,column=1,value='Q [W]')
    sheet.cell(row=1,column=2,value='Q_n [W]')
    sheet.cell(row=1,column=3,value='m_dot [kg/h]')
    sheet.cell(row=1,column=4,value='Twi [°C]')
    sheet.cell(row=1,column=5,value='Two [°C]')
    
    
    fillRange(sheet, 'A1','E1', 'FFDCE6F1')
    
    for iHX in range(0,n):                
        sheet.cell(row=2+iHX,column=1,value=Q_opt_sc[iHX])
        sheet.cell(row=2+iHX,column=2,value=q_n_sc[iHX])
        sheet.cell(row=2+iHX,column=3,value=m_dot_opt[iHX]*3600)
        if iHX > 0:
            sheet.cell(row=2+iHX,column=4,value=Twi_vec_opt[iHX-1])
        else:
            sheet.cell(row=2+iHX,column=4,value=problem_setup['Twi'])
        sheet.cell(row=2+iHX,column=5,value=Two_vec_opt[iHX])
        
    
    sheet.cell(row=iHX+4,column=3,value='Tbi [°C]:')    
    sheet.cell(row=iHX+4,column=4,value=Twi_vec_opt[iHX])
    sheet.cell(row=iHX+4,column=1,value='m_dot_p [kg/h]:')
    sheet.cell(row=iHX+4,column=2,value=m_dot_p_opt*3600)
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = 'Optimal solution -Heat flow [W]'
    chart1.y_axis.title = '[W]'
    
    data = Reference(sheet, min_col=1, min_row=1,\
                     max_row=1+n, max_col=2)
  
    chart1.add_data(data, titles_from_data=True)
 
    chart1.shape = 4
    chart1.height = 6
    chart1.width = 2*(n+2)
    
    chart1.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    chart1.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    sheet.add_chart(chart1, "G2")
    
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Optimal solution - Temperatures "
    chart2.y_axis.title = '[degC]'
    
    data = Reference(sheet, min_col=4, min_row=1,\
                     max_row=1 + n, max_col=5)

    chart2.add_data(data, titles_from_data=True)

    chart2.shape = 4
    chart2.height = 6
    chart2.width = 2*(n+2)
    
    chart2.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    chart2.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    sheet.add_chart(chart2, "G14")
else:
    sheet.merge_cells('B2:L2')
    sheet['B2'].alignment = Alignment(wrapText=True)
    sheet.row_dimensions[2].height = 40
    sheet.cell(row = 2, column = 2, value = 'Solver was not able to find any solution. Please, try to modify required heat flows or to choose a different catalogue list. To help to improve this tool, please contact me at ondrej.zlevor@cvut.cz. ')


    

# ------------------------------------------------------------------------
# ----------------------- Write selected results to sheet -------------------------
# ------------------------------------------------------------------------

    
sh_names = wb.get_sheet_names()
if any("Selected_results" == s for s in sh_names):  
    sheet = wb.get_sheet_by_name('Selected_results')
    wb.remove_sheet(sheet)
sheet = wb.create_sheet('Selected_results')

if not any(Q_res[0] == 0):
    iChosen = [0]
else:
    iChosen = []
if not any(Q_res[len(Q_res)-1] == 0):
    iChosen.append(len(Q_res)-1)
 
# find set with lowest Two temperature
Twi_min = problem_setup['Twi']
Twi_min_idx = 0
for iSet in range(1,len(Twi_all)):
    if Twi_all[iSet][n-1] < Twi_min and Twi_all[iSet][n-1] > 0 and final_obj[iSet] < 0.2:
        Twi_min = Twi_all[iSet][n-1]
        Twi_min_idx = iSet

if Twi_min < problem_setup['Twi']:
    iChosen.append(Twi_min_idx)
iChosen = list(set(iChosen))      # remove duplicit values

sheet.column_dimensions["F"].width = 30

if len(iChosen) == 0:
    sheet.merge_cells('B2:L2')
    sheet['B2'].alignment = Alignment(wrapText=True)
    sheet.row_dimensions[2].height = 40
    sheet.cell(row = 2, column = 2, value = 'Solver was not able to find any solution. Please, try to modify required heat flows or to choose a different catalogue list. To help to improve this tool, please contact me at ondrej.zlevor@cvut.cz. ')

for iSet in range(0,len(iChosen)):
    
    
    sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=1,value='Fobj')
    sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=2,value=final_obj[iChosen[iSet]])    
    sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=4,value='m_dot_p [kg/h]')
    sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=5,value=m_dot_all[iChosen[iSet]][n]*3600)
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=1,value='Q [W]')
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=2,value='Q_n [W]')
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=3,value='m_dot [kg/h]')
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=4,value='Twi [°C]')
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=5,value='Two [°C]')
    sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=6,value='Type')
    
    fillRange(sheet, 'A'+str(INIT_ROWS + iSet*n + iSet*varSpace),\
              'F'+str(INIT_ROWS + iSet*n + iSet*varSpace), 'FFDCE6F1')
    
    for iHX in range(0,n):
                    
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=1,value=Q_res[iChosen[iSet]][iHX])
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=1).number_format = '####'
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=2,value=q_n_all[iChosen[iSet]][iHX])
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=3,value=m_dot_all[iChosen[iSet]][iHX]*3600)
        if iHX > 0:
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=4,value=Twi_all[iChosen[iSet]][iHX-1])
        else:
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=4,value=problem_setup['Twi'])
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=5,value=Two_all[iChosen[iSet]][iHX])
        
        sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=6,value =\
                   q_n_names[q_n_unsrt.tolist().index(q_n_all[iChosen[iSet]][iHX])] )
    
    sheet.cell(row=iHX+INIT_ROWS+2 + iSet*n + iSet*varSpace,column=3,value='Tbi [°C]:')       
    sheet.cell(row=iHX+INIT_ROWS+2 + iSet*n + iSet*varSpace,column=4,value=Twi_all[iChosen[iSet]][iHX])
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    simName = ''.join([str(st) for st in q_heap[iChosen[iSet]][:]])
    simName = simName.replace('0', 'H')
    simName = simName.replace('1', 'L')
    chart1.title = "Set " + simName
    chart1.y_axis.title = '[W]'
    
    data = Reference(sheet, min_col=1, min_row=INIT_ROWS + iSet*n + iSet*varSpace,\
                     max_row=INIT_ROWS + iSet*n + iSet*varSpace + n, max_col=2)

    chart1.add_data(data, titles_from_data=True)

    chart1.shape = 4
    chart1.height = (n+varSpace)/2
    chart1.width = max(n+4,8)
    
    chart1.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    chart1.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    sheet.add_chart(chart1, "H" + str(1 + iSet*n + iSet*varSpace))
    
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Temperatures " + simName
    chart2.y_axis.title = '[degC]'
    
    data = Reference(sheet, min_col=4, min_row=INIT_ROWS + iSet*n + iSet*varSpace,\
                     max_row=INIT_ROWS + iSet*n + iSet*varSpace + n, max_col=5)

    chart2.add_data(data, titles_from_data=True)

    chart2.shape = 4
    chart2.height = (n+varSpace)/2
    chart2.width = max(n+4,8)
    
    chart2.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    chart2.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
    sheet.add_chart(chart2, sheet.cell(row =1 + iSet*n + iSet*varSpace,\
                                       column = 8 + int((n+4)/1.6) + 1).coordinate)

    


# ------------------------------------------------------------------------
# ----------------------- Write results to sheet -------------------------
# ------------------------------------------------------------------------

    
sh_names = wb.get_sheet_names()
if any("Particular_results" == s for s in sh_names):  
    sheet = wb.get_sheet_by_name('Particular_results')
    wb.remove_sheet(sheet)
sheet = wb.create_sheet('Particular_results')
   
sheet.column_dimensions["F"].width = 30

iSet = -1


for iSetHX in range(0,len(q_heap)):
    if not any(Q_res[iSetHX] == 0):
        iSet = iSet + 1
        
        sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=1,value='Fobj')
        sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=2,value=final_obj[iSetHX])
        sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=4,value='m_dot_p [kg/h]')
        sheet.cell(row=INIT_ROWS - 1 + iSet*n + iSet*varSpace,column=5,value=m_dot_all[iSetHX][n]*3600)
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=1,value='Q [W]')
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=2,value='Q_n [W]')
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=3,value='m_dot [kg/h]')
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=4,value='Twi [°C]')
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=5,value='Two [°C]')
        sheet.cell(row=INIT_ROWS + iSet*n + iSet*varSpace,column=6,value='Type')
        fillRange(sheet, 'A'+str(INIT_ROWS + iSet*n + iSet*varSpace),\
                  'F'+str(INIT_ROWS + iSet*n + iSet*varSpace), 'FFDCE6F1')
        for iHX in range(0,n):
                        
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=1,value=Q_res[iSetHX][iHX])
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=1).number_format= '####'
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=2,value=q_n_all[iSetHX][iHX])
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=3,value=m_dot_all[iSetHX][iHX]*3600)
            if iHX > 0:
                sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=4,value=Twi_all[iSetHX][iHX-1])
            else:
                sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=4,value=problem_setup['Twi'])
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=5,value=Two_all[iSetHX][iHX])
            
            sheet.cell(row=iHX+INIT_ROWS+1 + iSet*n + iSet*varSpace,column=6,value = q_n_names[q_n_unsrt.tolist().index(q_n_all[iSetHX][iHX])] )
        
        sheet.cell(row=iHX+INIT_ROWS+2 + iSet*n + iSet*varSpace,column=3,value='Tbi [°C]:')       
        sheet.cell(row=iHX+INIT_ROWS+2 + iSet*n + iSet*varSpace,column=4,value=Twi_all[iSetHX][iHX])
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        simName = ''.join([str(st) for st in q_heap[iSetHX][:]])
        simName = simName.replace('0', 'H')
        simName = simName.replace('1', 'L')
        chart1.title = "Set " + simName
        chart1.y_axis.title = '[W]'
        
        data = Reference(sheet, min_col=1, min_row=INIT_ROWS + iSet*n + iSet*varSpace,\
                         max_row=INIT_ROWS + iSet*n + iSet*varSpace + n, max_col=2)

        chart1.add_data(data, titles_from_data=True)

        chart1.shape = 4
        chart1.height = (n+varSpace)/2
        chart1.width = max(n+4,8)
        
        chart1.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
        chart1.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
        sheet.add_chart(chart1, "H" + str(1 + iSet*n + iSet*varSpace))
        
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Temperatures " + simName
        chart2.y_axis.title = '[degC]'
        
        data = Reference(sheet, min_col=4, min_row=INIT_ROWS + iSet*n + iSet*varSpace,\
                         max_row=INIT_ROWS + iSet*n + iSet*varSpace + n, max_col=5)
  
        chart2.add_data(data, titles_from_data=True)

        chart2.shape = 4
        chart2.height = (n+varSpace)/2
        chart2.width = max(n+4,8)
        
        chart2.y_axis.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
        chart2.title.tx.rich.p[0].r[0].rPr = CharacterProperties(sz=900)
        sheet.add_chart(chart2, sheet.cell(row =1 + iSet*n + iSet*varSpace,\
                                           column = 8 + int((n+4)/1.6) + 1).coordinate)
    
if iSet < 0:
    sheet.merge_cells('B2:L2')
    sheet['B2'].alignment = Alignment(wrapText=True)
    sheet.row_dimensions[2].height = 40
    sheet.cell(row = 2, column = 2, value = 'Solver was not able to find any solution. Please, try to modify required heat flows or to choose a different catalogue list. To help to improve this tool, please contact me at ondrej.zlevor@cvut.cz. ')

    
    
wb.save(xls_name_res)
os.startfile(xls_name_res)




